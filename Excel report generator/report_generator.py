import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel_report(cleaned_orders_file, output_file):
    """
    Create professional excel report from cleaned orders
    
    Args:
        cleaned_orders_file: CSV file with cleaned orders data
        output_file: Name of Excel file to create
    """

    #Read cleaned orders
    df = pd.read_csv(cleaned_orders_file, encoding='utf-8')

    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        #Sheet 1 : All Orders
        df.to_excel(writer, sheet_name='All Orders', index=False)

        #Sheet 2 : Summary
        summary = pd.DataFrame({
            'Metric' : [
                'Total Orders',
                'Valid Orders',
                'Invalid Orders',
                'Total Amount (¥)',
                'Average Order Value (¥)',
                'Highest Order (¥)',
                'Report Generated'
            ],
            'Value' : [
                len(df),
                len(df[df['is_valid'] == True]),
                len(df[df['is_valid'] == False]),
                df[df['is_valid'] == True]['amount'].sum(),
                df[df['is_valid'] == True]['amount'].mean(),
                df[df['is_valid'] == True]['amount'].max(),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        })
        summary.to_excel(writer, sheet_name='summary', index=False)

        # Sheet 3: Daily Breakdown
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            daily = df[df['is_valid'] == True].groupby(df['date'].dt.date).agg({
                'amount': ['count', 'sum']
            })
            daily.columns = ['Order Count', 'Total Amount']
            daily.to_excel(writer, sheet_name='Daily Breakdown')

                    # Sheet 4: Invalid Orders (for review)
        invalid = df[df['is_valid'] == False]
        if len(invalid) > 0:
            invalid.to_excel(writer, sheet_name='Invalid Orders - Review', index=False)

    # Format the Excel file
    format_excel_file(output_file)

    print(f"✅ Report generated successfully: {output_file}")

def format_excel_file(file_path):
    """Add professional formatting to excel file"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path)

    #Format summary sheet
    if 'summary' in wb.sheetnames:
        ws = wb['summary']

        #Bold Headers
        for cell in ws[1] :
            cell.font = Font(bold=True)

        #Add color to total amount
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == 'Total Amount (¥)':
                row[1].font = Font(bold=True, color='006600')  # Green color
                row[1].fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # Light green background

                wb.save('business_report.xlsx')

#Example usage
if __name__ == "__main__":
    generate_excel_report('cleaned_orders.csv', 'business_report.xlsx')
    print("📊 Report ready! Check business_report.xlsx")

